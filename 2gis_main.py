import random
import time
import openpyxl
from playwright.sync_api import sync_playwright, Page
from translate import Translator
from typing import List
from openpyxl import Workbook

class TwoGisMapParse:
    def __init__(self, keyword: str, sity: str, max_num_firm: int):
        self.keyword = keyword  #  Ищем по ключевому слову
        self.sity = sity  # Ищем в определённом городе
        self.max_num_firm = max_num_firm  # Максимальное количество фирм
        self.data_saving = '2gis_parse_results/data.xlsx'
        self.list_of_companies = []  # сюда добавляем списки из __get_firm_data, чтобы потом ввести их в xlsx

    def eng_sity(self):
        """Переводим город на английский для удобства"""
        return self.translate_text(self.sity).lower()

    def __get_links(self) -> List[str]:
        """Извлекаем ссылки на организации, находящиеся на странице"""
        print("Собираем ссылки на организации с текущей страницы...")

        link_selector = 'a[href*="/firm/"]'
        
        found_links = self.page.query_selector_all(link_selector)  # Ищем только видимые карточки организаций(firm)
        for count, link in enumerate(found_links):
            if len(self.list_of_companies)  >= self.max_num_firm:  # Делаем так, чтобы кол-во не превышало желамое кол-во объявлений
                break 
            if not link.is_visible():  # Проверяем, видим ли элемент
                continue
            href = (link.get_attribute("href") or "")  # Находим элемент на стр., где есть /firm/
            # На всякий случай делаю ещё проверку; Ещё проверяю город, чтоб не искало в регионах
            if (href and "/firm/" in href and self.eng_sity() in href):
                href = f"https://2gis.ru{href}"  # Делаем полное url
                firm_data = self.__get_firm_data(url=href)  # Ищем все данные фирмы
                self.list_of_companies.append(firm_data)  # Добавляем в список, который потом пойдет в xlsx

    def __get_firm_data(self, url: str):
        """Берем данные фирмы: название, телефон, сайт"""
        self.page2 = self.context.new_page()  # Создаем новую страницу
        self.page2.goto(url=url)  # Переходим на неё
        
        true_phone = "Телефон не найден"  # Если будет не найдено
        true_site = "Нет ссылки на сайт"
        
        # Название фирмы
        firm_title = self.page2.title().split(',')[0]  #Отделяем: (Назв.фирмы, ул. ...)
        
        # Номер телефона
        try:
            # Находим контейнер, затем ссылку внутри него
            phone_container = self.page2.query_selector(':has(button:has-text("Показать телефон")):has(a[href^="tel:"])')
            if phone_container:
                # Теперь ищем телефон внутри этого контейнера
                phone = phone_container.query_selector('a[href^="tel:"]')
                true_phone = phone.get_attribute("href")[4:]  # Вывожу без tel:
            else:
                true_phone = "Контейнер с телефоном и кнопкой не найден"
        except Exception as e:
            true_phone = f"Ошибка: {e}"
        
        # Название сайта
        site_elements = self.page2.query_selector_all('a[href^="https://link.2gis.ru/"]')  # Ищем ссылки(сайт)
        if site_elements:  # Если есть хоть одна ссылка
            site_texts = [element.text_content().strip() for element in site_elements]
            try:
                a = list(filter(lambda i: i if ('.ru' in i or '.com' in i or '.net' in i or '.рф' in i) and
                            '@' not in i else '', site_texts))  # Фильтруем, чтоб выводилось нужное 
                true_site = f"{a[0]}"
            except:
                true_site = "Нет ссылки на сайт"
                
        self.page2.close()
        return [url, firm_title, true_phone, true_site, '-']

    def data_output_to_xlsx(self, get_firm_data):
        '''Выводим данные в файл xlsx'''
        # Создать новый файл (старый перезапишется)
        wb = Workbook()
        ws = wb.active
        
        # Добавить заголовки
        headers = ["URL", "Название", "Телефон", "Сайт"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Начинаем запись со второй строки (после заголовков)
        start_row = 2

        # Цикл по данным фирм
        for firm_data in get_firm_data:  # firm_data - это список ['URL', 'Название', 'Телефон', 'Сайт']
            # Запись каждой строки в Excel
            for col, value in enumerate(firm_data, start=1):
                ws.cell(row=start_row, column=col, value=value)
            start_row += 1  # Перейти на следующую строку

        # Сохранить файл
        wb.save(self.data_saving)
        print(f"Записано {len(get_firm_data)} строк в файл data.xlsx")

    def translate_text(self, text, from_lang="ru", to_lang="en"):
        # Создаем объект Translator, указывая исходный язык и язык перевода
        translator = Translator(from_lang=from_lang, to_lang=to_lang)
        try:
            # Пытаемся перевести текст
            translated_text = translator.translate(text)
            return translated_text  # Возвращаем переведенный текст
        except Exception as e:
            # Если возникает ошибка, возвращаем сообщение об ошибке
            return f"Error: {e}"

    def get_random_user_agent(self):
        ''''''
        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
        ]
        return random.choice(user_agents)

    def parse(self):
        """Парсинг сайта"""
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=False)  # headless=False - без графического итерфейса
            self.context = browser.new_context(
                user_agent=self.get_random_user_agent()
                )  # По типу вкладок инкогнито
            self.page = self.context.new_page()  # Новая страница, создается в контексте
            self.page.goto(f"https://2gis.ru/{self.eng_sity()}")  #  Переходим по адресу с переведенным городом
            # Ищем поле поиска, пишем туда keyword и печатаем каждую букву с промежутком времени 0.4 с
            self.page.get_by_placeholder("Поиск в 2ГИС").type(text=self.keyword, delay=0.4)
            self.page.keyboard.press("Enter")  # Нажимаем Enter
            time.sleep(3)  # Задержка для загрузки страницы
            self.__get_links()  # Вывод ссылок на организации 
            self.page.click('[style="transform: rotate(-90deg);"]')
            time.sleep(3)  # Задержка для загрузки страницы
            self.__get_links()  # Вывод ссылок на организации
            self.data_output_to_xlsx(self.list_of_companies)
            # self.page.click('[style="transform: rotate(-90deg);"]')  # Кликаем на кнопку перехода на след. страницу
            # time.sleep(4)
            # self.__get_links()
            time.sleep(180)


if __name__ == "__main__":
    TwoGisMapParse("Мойка", "Самара", 22).parse()  # Ключевое слово, город, кол-во объявлений
