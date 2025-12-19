import random
import os
import openpyxl
import asyncio
from playwright.async_api import async_playwright
from typing import List
from openpyxl import Workbook
from googletrans import Translator
from collections import deque


class TwoGisMapParse:
    def __init__(self, keyword: str, sity: str, max_num_firm: int):
        self.keyword = keyword  # Ищем по ключевому слову
        self.sity = sity  # Ищем в определённом городе
        self.max_num_firm = max_num_firm  # Максимальное количество фирм
        self.data_saving = "2gis_parse_results/data.xlsx"
        self.list_of_companies = deque()  # сюда добавляем списки из __get_firm_data, чтобы потом ввести их в xlsx
        self.start_row = 2
        if os.path.exists(self.data_saving):
            os.remove(self.data_saving)

    async def random_delay(self, min_seconds=1, max_seconds=3):
        """Случайная задержка между действиями"""
        await asyncio.sleep(random.uniform(min_seconds, max_seconds))

    async def translate_text(self):
        """Переводим город на английский для удобства"""
        self.translator = Translator()
        a = await self.translator.translate(self.sity, src="ru", dest="en")
        a = '-'.join(a.text.split())
        return a.lower()

    async def __get_links(self) -> List[str]:
        """Извлекаем ссылки на организации, находящиеся на странице"""
        print("Собираем ссылки на организации с текущей страницы...")
        self.list_of_companies = []
        link_selector = 'a[href*="/firm/"]'

        found_links = await self.page.query_selector_all(link_selector)  # Ищем только видимые карточки организаций(firm)
        for link in found_links:
            if not await link.is_visible():  # Проверяем, видим ли элемент
                continue
            href = await link.get_attribute("href") or ""  # Находим элемент на стр., где есть /firm/
            # На всякий случай делаю ещё проверку; Ещё проверяю город, чтоб не искало в регионах
            if href and "/firm/" in href and await self.translate_text() in href:
                href = f"https://2gis.ru{href}"  # Делаем полное url
                if self.ws.max_row + len(self.list_of_companies) - 1>= self.max_num_firm:
                    break
                firm_data = await self.__get_firm_data(url=href)  # Ищем все данные фирмы

                if self.true_phone != "---" or (self.true_phone == "---" and self.true_site != "Нет ссылки на сайт"):
                    self.list_of_companies.append(firm_data)  # Добавляем в список, который потом пойдет в xlsx

    async def __get_firm_data(self, url: str):
        """Берем данные фирмы: название, телефон, сайт"""
        self.page2 = await self.context.new_page()  # Создаем новую страницу
        await self.page2.goto(url=url)  # Переходим на неё

        self.true_phone = "Телефон не найден"  # Если будет не найдено
        self.true_site = "Нет ссылки на сайт"

        # Название фирмы
        firm_title = (await self.page2.title()).split(",")[0]  # Отделяем: (Назв.фирмы, ул. ...)
        # Категория
        firm_category = (await self.page2.title()).split(",")[1]  # Отделяем: (Назв.фирмы, ул. ...)

        # Номер телефона
        try:
            # Находим контейнер, затем ссылку внутри него
            phone_container = await self.page2.query_selector(':has(button:has-text("Показать телефон")):has(a[href^="tel:"])')
            if phone_container:
                # Теперь ищем телефон внутри этого контейнера
                phone = await phone_container.query_selector('a[href^="tel:"]')
                self.true_phone = (await phone.get_attribute("href"))[4:16]  # Вывожу без tel:
            else:
                self.true_phone = "---"
        except Exception as e:
            self.true_phone = f"Ошибка {e}"

        # Название сайта
        site_elements = await self.page2.query_selector_all('a[href^="https://link.2gis.ru/"]')  # Ищем ссылки(сайт)
        if site_elements:  # Если есть хоть одна ссылка
            site_texts = [await element.text_content() for element in site_elements]
            try:
                a = list(filter(lambda i: (i if (".ru" in i or ".com" in i or 
                                                 ".рф" in i or ".net" in i
                                                 ) and "@" not in i else ""),site_texts,))  # Фильтруем, чтоб выводилось нужное
                self.true_site = f"{a[0]}"
            except:
                self.true_site = "Нет ссылки на сайт"
        await self.page2.close()
        return ['', firm_title, firm_category, self.true_phone, self.true_site, "-"]

    async def check_xlsx(self):
        """Функция для создания заготовки под xlsx файл"""
        # Проверки, есть ли папка, если нет, то создаем
        os.makedirs("2gis_parse_results", exist_ok=True)

        # Создать новый файл (старый удаляется при включении программы)
        self.wb = Workbook()
        self.ws = self.wb.active

        # Добавляем заголовки
        headers = ["URL", "Название", "Категория", "Телефон", "Сайт"]
        for col, header in enumerate(headers, start=1):
            self.ws.cell(row=1, column=col, value=header)

    async def data_output_to_xlsx(self, get_firm_data):
        """Выводим данные в файл xlsx"""
        # Открыть существующий файл
        if os.path.exists(self.data_saving):
            self.wb = openpyxl.load_workbook(self.data_saving)
            self.ws = self.wb.active
        # Цикл по данным фирм
        for firm_data in get_firm_data:  # firm_data - это список ['URL', 'Название', 'Телефон', 'Сайт']
            # Запись каждой строки в Excel
            for col, value in enumerate(firm_data, start=1):
                self.ws.cell(row=self.start_row, column=col, value=value)
            self.start_row += 1  # Перейти на следующую строку

        # Сохранить файл
        self.wb.save(self.data_saving)
        print(self.list_of_companies)
        print(f"Записано {len(get_firm_data)} строк в файл data.xlsx")

    async def get_random_user_agent(self):
        """Скрываем автоматизацию с помощью захода с разных систем"""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
        ]
        return random.choice(user_agents)

    async def parse_main(self, update_callback=None):
        """Парсинг сайта"""
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=False)  # headless=False - без графического итерфейса
            self.context = await browser.new_context(
                user_agent=await self.get_random_user_agent(),
                locale="ru-RU",
                timezone_id="Europe/Moscow",
            )  # По типу вкладок инкогнито
            self.page = (await self.context.new_page())  # Новая страница, создается в контексте
            await self.page.goto(f"https://2gis.ru/{await self.translate_text()}", wait_until="domcontentloaded")  # Переходим по адресу с переведенным городом
            if await self.translate_text() not in self.page.url:
                await self.page.close()
            # Ищем поле поиска, пишем туда keyword и печатаем каждую букву с промежутком времени 0.4 с
            await self.page.get_by_placeholder("Поиск в 2ГИС").type(text=self.keyword, delay=0.4)
            await self.page.keyboard.press("Enter")  # Нажимаем Enter
            await self.random_delay(3, 4)  # Задержка для загрузки страницы
            await self.check_xlsx()
            # Собираем данные с задержками
            while self.ws.max_row < self.max_num_firm:
                print(self.ws.max_row)
                await self.__get_links()  # Ищем ссылки и данные организаций
                await self.data_output_to_xlsx(self.list_of_companies)  # Записываем данные в Excel
                # Имитация просмотра страницы
                await self.random_delay(1, 2)

                # Переход на следующую страницу с проверкой
                next_button = await self.page.query_selector('[style="transform: rotate(-90deg);"]')
                if next_button and await next_button.is_visible():
                    await self.random_delay(1, 2)
                    await next_button.click()
                    await self.random_delay(3, 4.5)  # Ждем загрузки следующей страницы
                else:
                    break  # Больше нет страниц
            else:
                await self.page.close()

            print(f"Записано {self.ws.max_row - 1} организаций")
            await asyncio.sleep(4)
            self.page.close()


async def main():
    parser = TwoGisMapParse(keyword="Авто Мойка", sity="Саратов", max_num_firm=200)
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())
